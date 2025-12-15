try:
    import openpyxl
    
    # ì—‘ì…€ íŒŒì¼ ì—´ê¸°
    wb = openpyxl.load_workbook('guinness_test_data.xlsx')
    
    print("=" * 60)
    print("ğŸ“Š guinness_test_data.xlsx íŒŒì¼ ì •ë³´")
    print("=" * 60)
    
    print(f"\nì‹œíŠ¸ ì´ë¦„: {wb.sheetnames}")
    
    # ì²« ë²ˆì§¸ ì‹œíŠ¸ ì„ íƒ
    ws = wb.active
    
    # ëª¨ë“  ë°ì´í„° ì½ê¸°
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    print(f"\nì´ í–‰ ìˆ˜: {len(data)}ê°œ")
    
    if len(data) > 0:
        print(f"\nì»¬ëŸ¼ëª… (ì²« í–‰): {data[0]}")
        
        print("\nì²« 10ê°œ ë°ì´í„°:")
        for i, row in enumerate(data[:11], 0):
            if i == 0:
                print(f"í—¤ë”: {row}")
            else:
                print(f"{i}: {row}")
    
    wb.close()
    
except ImportError:
    print("openpyxlì´ ì„¤ì¹˜ë˜ì–´ ìˆì§€ ì•ŠìŠµë‹ˆë‹¤.")
    print("ì„¤ì¹˜ ëª…ë ¹: pip install openpyxl")
except Exception as e:
    print(f"ì˜¤ë¥˜ ë°œìƒ: {e}")
    import traceback
    traceback.print_exc()

