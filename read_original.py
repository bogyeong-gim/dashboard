try:
    import openpyxl
    
    # 원본 엑셀 파일 열기
    wb = openpyxl.load_workbook('guinness_test_data (1).xlsx')
    
    print("=" * 70)
    print("📊 guinness_test_data (1).xlsx 파일 정보 (원본)")
    print("=" * 70)
    
    print(f"\n시트 이름: {wb.sheetnames}")
    
    # 첫 번째 시트 선택
    ws = wb.active
    
    # 모든 데이터 읽기
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    print(f"\n총 행 수: {len(data)}개")
    
    if len(data) > 0:
        print(f"\n컬럼명 (첫 행): {data[0]}")
        
        print("\n모든 데이터:")
        for i, row in enumerate(data, 0):
            if i == 0:
                print(f"헤더: {row}")
            else:
                print(f"{i}: {row}")
    
    wb.close()
    
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다.")
    print("설치 명령: pip install openpyxl")
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()

